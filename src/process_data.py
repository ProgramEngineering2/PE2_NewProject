import os
import pandas as pd
from src.flat_transmission import plot_flat_transmission_spectra
from src.ivcurve import plot_iv_data
from src.pandas_frame import pandas_data, save_to_excel
from src.ref_transmission import plot_transmission_spectra
from src.transmission import plot_transmission_spectra_all
from src.device_waferno_find_xml import find_xml_files
import xml.etree.ElementTree as ET
import matplotlib.pyplot as plt
import numpy as np
import warnings
import openpyxl


# RankWarning 경고 무시
warnings.simplefilter('ignore', np.RankWarning)
warnings.filterwarnings("ignore", "No artists with labels found to put in legend.  Note that artists whose label start with an underscore are ignored when legend() is called with no argument.", UserWarning)

def process_data(device, wafer_nos):

    directories = [
        'dat/HY202103/D07/20190715_190855',
        'dat/HY202103/D08/20190526_082853',
        'dat/HY202103/D08/20190528_001012',
        'dat/HY202103/D08/20190712_113254',
        'dat/HY202103/D23/20190528_101900',
        'dat/HY202103/D23/20190531_072042',
        'dat/HY202103/D23/20190603_204847',
        'dat/HY202103/D24/20190528_105459',
        'dat/HY202103/D24/20190528_111731',
        'dat/HY202103/D24/20190531_151815',
        'dat/HY202103/D24/20190603_225101'
    ]

    # Excel 파일 저장
    from datetime import datetime
    # 현재 날짜와 시간 불러오기
    now = datetime.now()
    formatted_now = now.strftime("%Y%m%d_%H%M")

    excel_directory = f'res/CSV/{formatted_now}' #엑셀파일이 들어가있는 디렉토리까지의 경로
    final_df = pandas_data(device, wafer_nos)
    save_to_excel(final_df, excel_directory, f'Analysis_Data_{formatted_now}.xlsx') #pandas data의 최종 결과 데이터 프레임을 엑셀 디렉토리 경로를 따라 filename으로 저장.
    wb = openpyxl.load_workbook(os.path.join(excel_directory, f'Analysis_Data_{formatted_now}.xlsx'))
    ws = wb.active

    # XML 파일 찾기
    xml_files = find_xml_files(directories, device, wafer_nos)
    if not xml_files:
        print("XML 파일을 찾을 수 없습니다.")
        return

    image_files = []
    for i in range(len(xml_files)):
        image_files.append('link')

    # 권한 확인 함수
    def check_permissions(file_path):
        permissions = {
            'exists': os.path.exists(file_path),
            'readable': os.access(file_path, os.R_OK),
            'writable': os.access(file_path, os.W_OK),
            'executable': os.access(file_path, os.X_OK)
        }
        return permissions

    jpgs_directory = os.path.join('res', 'jpgs')

    # 하이퍼링크 추가
    for row, jpg_file in zip(range(2, len(xml_files) + 2), image_files):
        Lot = ws[f'A{row}'].value
        Wafer = ws[f'B{row}'].value
        Mask = ws[f'C{row}'].value
        TestSite = ws[f'D{row}'].value
        Row = ws[f'I{row}'].value
        Column = ws[f'J{row}'].value
        filename = f'{Lot}_{Wafer}_({Row},{Column})_{Mask}_{TestSite}.xml'
        wafer_id = filename[9:12]
        no_xml_name = os.path.splitext(filename)[0]
        jpg_filename = os.path.splitext(filename)[0] + '.jpg'
        cell = ws.cell(row=row, column=19)  # 두 번째 열에 하이퍼링크 추가
        jpg_file = jpg_filename  # 이미지 파일명 예시
        file_path = os.path.join(jpgs_directory, wafer_id, formatted_now, jpg_filename)
        absolute_path = os.path.abspath(file_path)
        cell.hyperlink = absolute_path
        cell.value = 'link'  # 하이퍼링크 표시 텍스트
        cell.style = "Hyperlink"

    # 엑셀 파일 저장
    wb.save(os.path.join(excel_directory, f'Analysis_Data_{formatted_now}.xlsx'))

    # XML 파일 찾기
    xml_files = find_xml_files(directories, device, wafer_nos)
    if not xml_files:
        print("XML 파일을 찾을 수 없습니다.")
        return

    # 그래프 생성 및 저장
    jpgs_directory = os.path.join('res', 'jpgs')
    for xml_file in xml_files:
        fig, axs = plt.subplots(2, 2, figsize=(15, 10))

        tree = ET.parse(xml_file)
        root = tree.getroot()

        plot_iv_data(axs[0, 0], root)
        plot_transmission_spectra(axs[0, 1], root)
        plot_transmission_spectra_all(axs[1, 0], root)
        plot_flat_transmission_spectra(axs[1, 1], root)


        # subplot 간의 간격 조정
        plt.subplots_adjust(wspace=0.4, hspace=0.4)

        # 파일명에서 9번~12번 글자 추출
        filename = os.path.basename(xml_file)
        wafer_id = filename[9:12]

        # 파일 경로 설정
        jpgs_wafer_directory = os.path.join(jpgs_directory, wafer_id,formatted_now)
        if not os.path.exists(jpgs_wafer_directory):
            os.makedirs(jpgs_wafer_directory)

        # JPG 파일로 저장
        jpg_filename = os.path.splitext(filename)[0] + '.jpg'
        plt.savefig(os.path.join(jpgs_wafer_directory, jpg_filename))
        plt.close()  # 그래프 초기화